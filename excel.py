# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color
import os
import re
import string

wb = openpyxl.Workbook()
s1 = wb.active
s1.title = ("표지")
s2 = wb.create_sheet("점검대상")
#s3 = wb.create_sheet("Unix서버요약")


################ 표지 작성 ################ 
s1.merge_cells("B17:J17")
s1['B17'] = "UNIX 서버 취약점 점검 상세 보고서"
s1['B17'].font = Font(size=26, bold=True)
s1['B17'].alignment = Alignment(horizontal='center', vertical='center')
s1.row_dimensions[17].height = 66.75
for a in range(9):
    s1.cell(row=17, column=a + 2).border = Border(top=Side(style='double'), bottom=Side(style='double'))

s1.merge_cells("A31:K31")
s1['A31'] = "2021.05"
s1['A31'].font = Font(size=20, bold=True)
s1['A31'].alignment = Alignment(horizontal='center', vertical='center')
s1.row_dimensions[31].height = 41.25


################  점검 대상 작성 ################ 

s2.row_dimensions[1].height = 31.50
s2.row_dimensions[2].height = 6
s2.row_dimensions[3].height = 21.75
s2.column_dimensions['A'].width = 5.38
s2.column_dimensions['B'].width = 11.88
s2.column_dimensions['C'].width = 23.38
s2.column_dimensions['D'].width = 26.13
s2.column_dimensions['E'].width = 13.88
s2.column_dimensions['F'].width = 13.88

s2.merge_cells("A1:G1")
s2['A1'] = "점검 대상"
s2['A1'].font = Font(size=20, bold=True)
s2['A3'] = "No"
s2['B3'] = "구분"
s2['C3'] = "호스트명"
s2['D3'] = "OS"
s2['E3'] = "IP"
s2['F3'] = "용도"
s2['G3'] = "비고"

for a in range(7):
    s2.cell(row=3, column=a + 1).border = Border(left=Side(style='dotted'), 
                                                 top=Side(style='medium'), 
                                                 right=Side(style='dotted'), 
                                                 bottom=Side(style='medium'))
    s2.cell(row=3, column=a + 1).alignment = Alignment(horizontal='center', vertical='center')
    s2.cell(row=3, column=a + 1).fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')
    s2.cell(row=3, column=a + 1).font = Font(size=9, bold=True)

path = "server/"
i = 1
for p in os.listdir(path):
    s2.append([i, "Linux Server",p.split("_")[1]])
    for a in range(7):
        s2.cell(row=3 + i, column=a + 1).border = Border(left=Side(style='dotted'), 
                                                         right=Side(style='dotted'), 
                                                         bottom=Side(style='dotted'))
        s2.cell(row=3 + i, column=a + 1).alignment = Alignment(horizontal='center', vertical='center')
        s2.cell(row=3 + i, column=a + 1).font = Font(size=9)
        s2.row_dimensions[3 + i].height = 22.5
    i += 1

    
################ 서버 별 스크립트 분석 결과 작성 ################ 

checklist = ["root 계정 원격 접속 제한","패스워드 복잡성 설정","계정 잠금 임계값 설정","패스워드 파일 보호","root 이외의 UID가 '0' 금지","root 계정 su 제한","패스워드 최소 길이 설정","패스워드 최대 사용기간 설정","패스워드 최소 사용기간 설정","불필요한 계정 제거","관리자 그룹에 최소한의 계정 포함","계정이 존재하지 않는 GID 금지","동일한 UID 금지","사용자 shell 점검","Session Timeout 설정","root 홈, 패스 디렉토리 권한 및 패스 설정","파일 및 디렉토리 소유자 설정","/etc/passwd 파일 소유자 및 권한 설정","/etc/shadow 파일 소유자 및 권한 설정","/etc/hosts 파일 소유자 및 권한 설정","/etc/(x)inetd.conf 파일 소유자 및 권한 설정","/etc/syslog.conf 파일 소유자 및 권한 설정","/etc/services 파일 소유자 및 권한 설정","SUID, SGID, Sticky bit 설정 파일 점검","사용자, 시스템 시작파일 및 환경파일 소유자 및 권한 설정","world writable 파일 점검","/dev에 존재하지 않는 device 파일 점검","$HOME/.rhosts, hosts.equiv 사용 금지","접속 IP 및 포트 제한","hosts.lpd 파일 소유자 및 권한 설정","NIS 서비스 비활성화","UMASK 설정 관리","홈 디렉토리 소유자 및 권한 설정","홈 디렉토리로 지정한 디렉토리의 존재 및 관리","숨겨진 파일 및 디렉토리 검색 및 제거(dot file)","finger 서비스 비활성화","Anonymous ftp 비활성화","r 계열 서비스 비활성화","cron 파일 소유자 및 권한 설정","DoS 공격에 취약한 서비스 비활성화","NFS 서비스 비활성화","NFS 접근 통제","automountd 제거","RPC 서비스 확인","NIS, NIS+ 점검","tftp, talk 서비스 비활성화","sendmail 버전 점검","스팸 메일 릴레이 제한","일반 사용자의 sendmail 실행 방지","DNS 보안 패치","DNS Zone Transfer 설정","Apache 디렉토리 리스팅 제거","Apache 웹 프로세스 권한 제한","Apache 상위 디렉토리 접근 금지","Apache 불필요한 파일 제거","Apache 링크 사용 금지","Apache 파일 업로드 및 다운로드 제한","Apache 웹 서비스 영역의 분리","ssh 원격 접속 허용","ftp 서비스 확인","ftp 계정 shell 제한","ftpusers 파일 소유자 및 권한 설정","ftpusers 파일 설정","at 파일 소유자 및 권한 설정","SNMP 서비스 구동 점검","SNMP 서비스 커뮤니티 스트링의 복잡성 설정","로그온 시 경고 메시지 제공","NFS 설정 파일 접근 권한","expn, vrfy 명령어 제한","Apache 웹 서비스 정보 숨김","최신 보안 패치 및 벤더 권고사항 적용","로그의 정기적 검토 및 보고","정책에 따른 시스템 로깅 설정"]
important = ["상","상","상","상","중","하","중","중","중","하","하","하","중","하","하","상","상","상","상","상","상","상","상","상","상","상","상","상","상","하","중","중","중","중","하","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","상","중","하","중","중","중","중","중","중","하","중","중","중","상","상","하"]
result = []
for p in os.listdir(path):
    server_name = p.split("_")[1]
    s = wb.create_sheet(server_name)
    
    s.row_dimensions[6].height = 27.75
    s.column_dimensions['A'].width = 5
    s.column_dimensions['B'].width = 5
    s.column_dimensions['C'].width = 43.50
    s.column_dimensions['D'].width = 5.5
    s.column_dimensions['E'].width = 5.5
    s.column_dimensions['F'].width = 61.25
    s.column_dimensions['G'].width = 61.25
    s.column_dimensions['H'].width = 61.25

    s.merge_cells("A1:B1")
    s['A1'] = "HOSTNAME"
    s['C1'] = server_name
    s.merge_cells("A2:B2")
    s['A2'] = "OS"
    s.merge_cells("A3:B3")
    s['A3'] = "IP"
    s.merge_cells("A4:B4")
    s['A4'] = "용도"
    
    for a in range(4):
        for b in range (3):
            if (b < 2):
                s.cell(row=1 + a, column=b + 1).fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')
                s.cell(row=1 + a, column=b + 1).font = Font(size = 9, bold=True)
            else:
                s.cell(row=1 + a, column=b + 1).font = Font(size=9)
            s.cell(row=1 + a, column=b + 1).alignment = Alignment(horizontal='center', vertical='center')
            s.cell(row=1 + a, column=b + 1).border = Border(left=Side(style='dotted'), 
                                                            top=Side(style='dotted'),
                                                            right=Side(style='dotted'), 
                                                            bottom=Side(style='dotted'))
    
    for a in range(3):
        s.cell(row=4, column=a + 1).border = Border(bottom=Side(style='medium'), right=Side(style='dotted'))
    
    s.merge_cells("A6:B6")
    s['A6'] = "구분"
    s['C6'] = "점 검 항 목"
    s['D6'] = "항목\n중요도"
    s['E6'] = "점검\n결과"
    s['F6'] = "현재 상태"
    s['G6'] = "취약 이유"
    s['H6'] = "취약할 시 문제점"
    
    for a in range(8):
        s.cell(row=6, column=a + 1).border = Border(left=Side(style='dotted'), 
                                                    top=Side(style='medium'),
                                                    right=Side(style='dotted'), 
                                                    bottom=Side(style='medium'))
        s.cell(row=6, column=a + 1).fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')
        s.cell(row=6, column=a + 1).font = Font(size = 10, bold=True)
        s.cell(row=6, column=a + 1).alignment = Alignment(horizontal='center', vertical='center')
    
    
    print("################ [" + server_name + "] 점검 시작 ################ ")
    
    num = ""
    result = []
    subt = re.compile('^\[\s.* \]')
    comment = re.compile('\s*#.*')
    cn = re.compile("\d*/g")
    d = re.compile('\d+')
    f = open(path + p, 'r', encoding='UTF8')
    while True:
        line = f.readline()
        if not line: break
        
        if "[ U-" in line:
            num = line.split(" ")[1]
        
        ### root 계정 원격 접속 제한
        if "U-01" in num:
            n = 1
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "permitrootlogin" in line.lower():
                    if "yes" in line.lower():
                        state = "N"
                        explain += "PermitRootLogin 설정이 yes로 되어있다. \n이 설정을 제거/주석 처리하거나 no로 바꿔야 한다.\n" 
                        code += line
                    else:
                        if "PermitRootLogin" not in explain:
                            explain += "PermitRootLogin 설정이 올바르게 되어 있다.\n"
                        code += line
                elif "console" in line.lower():
                    flag = 1
                    while True:
                        l = f.readline()
                        if " " in l:
                            break;
                        
                        if "pts" in l.lower():
                            code += l
                            flag = 0
                    
                    if flag == 1:
                        explain += "pts 관련 설정이 존재하지 않는다.\n"
                    else:
                        state = "N"
                        explain += "pts 관련 설정이 존재하므로 취약함\n"
                elif "pam_securetty.so" in line.lower():
                    if "required" in line.lower() or (
                        "success=ok" in line.lower() and
                        #"new_authtok_reqd=ok" in line.lower() and
                        "ignore=ignore" in line.lower() and
                        "default=bad" in line.lower()):
                        explain +="pam_securetty.so 설정이 권고사항과 동일하게 설정되어 있다.\n"
                    else:
                        state = "N"
                        explain += "pam_securetty.so 설정에 취약점이 존재한다.\n"
                    
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 패스워드 복잡성 설정
        elif "U-02" in num:
            n = 2
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            flag = 0
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "password" in line:
                    code += line
                    flag = 1
                    if int(line.split("minlen=")[1][0]) >= 10:
                        if line.split("minlen=")[1][0].count("-1") >= 2:
                            state = "Y"
                            explain += "패스워드 최소 길이가 10자리 이상이고, 2가지 이상의 문자가 최소 입력 기능 설정되기 때문에 양호하다.\n"
                        else:
                            explain += "패스워드 최소 길이가 10자리 이상이지만, 2가지 미만의 문자가 최소 입력 기능 설정되어 취약\n"
                    elif int(line.split("minlen=")[1][0]) >= 8:
                        if line.split("minlen=")[1][0].count("-1") >= 3:
                            state = "Y"
                            explain += "패스워드 최소 길이가 8자리 이상이고, 3가지 이상의 문자가 최소 입력 기능 설정되기 때문에 양호하다.\n"
                        else:
                            explain += "패스워드 최소 길이가 8자리 이상이지만, 3가지 미만의 문자가 최소 입력 기능 설정되어 취약\n"
                    else:
                        state = "N"
                        explain += "패스워드 최소 길이가 8자리 미만이기 때문에 취약"
                elif "존재하지 않음"  in line:
                    code += line
                #elif "수동확인" in line and flag == 0:
                    #if state == "N":
                        #state = "C"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ### 계정 잠금 임계값 설정
        elif "U-03" in num:
            n = 3
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif " deny=" in line.lower() and "거부" not in line:
                    code += line
                    if int(line.lower().split("deny=")[1][0]) <= 10:
                        state = "Y"
                        explain += "계정 잠금 임계값이 10회 이하의 값으로 설정되어 있어 양호\n"
                    else:
                        state = "N"
                        explain += "계정 잠금 임계값이 10회 이상의 값으로 설정되어 있어 취약\n"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "계정 잠금 임계값이 설정되어 있지 않아 취약\n"
                state = 'N'
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 패스워드 파일 보호
        elif "U-04" in num:
            n = 4
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/etc/passwd 파일" in line:
                    line = f.readline()
                    code += line
                    
                    if line.split(":")[1] != "x":
                        state = "N"
                        explain += "패스워드를 암호화하여 저장하지 않는다.\n"
                    else:
                        state = "Y"
                        explain += "패스워드를 암호화하여 저장하였다.\n"
                    
                    line = f.readline()
                    explain += line
                    
                    if "사용하고 있습니다." not in line:
                        state = "N"
                    
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ### root 이외의 UID가 '0' 금지
        elif "U-05" in num:
            n = 5
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "GID 추가 0 확인" in line: 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "0" in line and "root" not in line:
                    state = "N"
                    explain += "root와 동일한 UID를 가진 계정이 존재함으로 취약\n"
                    code += line
                    
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/etc/passwd" in line:
                    continue
                elif "0" in line and line.split(" ")[0] not in ["daemon", "bin", "adm", "uucp", "nuucp", "lp", "hpdb", "sync", "shutdown", "halt", "operator", "root"]:
                    state = "Y"
                    explain += "root와 동일한 GID를 가진 계정이 존재한다.\n"
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "root와 동일한 UID/GID를 가진 계정이 존재하지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### SU 명령 사용제한
        elif "U-06" in num:
            n = 6
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "PAM 모듈 확인" in line: 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                if "/bin/su 실행 권한 및 소유그룹 확인(4750)" in line:
                    line = f.readline()
                    
                    if "-rwsr-x---" not in line:
                        state = "N"
                        explain+="/bin/su 파일의 권한이 적절하게 설정되어 있지 않다.\n"
                    else:
                        explain +="/bin/su 파일의 권한이 적절하게 설정되어 있다.\n"
            
            while True:
                if " END" in line: 
                    num = ""
                    break
                    
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "required" in line and "pam_wheel.so" in line:
                    code += line
                    if comment.match(line) == None:
                        explain += "PAM 모듈의 허용 그룹 설정이 되어 있다.\n"
                    else:
                        state = "N"
                        explain += "PAM 모듈의 허용 그룹 설정이 주석 처리 되어 있다.\n"
                
                line = f.readline()
                
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 패스워드 최소 길이 설정
        elif "U-07" in num:
            n = 7
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "PASS_MIN_LEN" in line and comment.match(line) == None:
                    code += line
                    if int(line.split("\t")[1][0]) < 8:
                        state = "N"
                elif "minlen=" in line:
                    code += line
                    if int(line.split("minlen=")[1][0]) < 8:
                        state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "패스워드 최소 길이가 8 이상이므로 양호\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "패스워드 최소 길이가 8 미만이므로 취약\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 패스워드 최대 사용기간 설정
        elif "U-08" in num:
            n = 8
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None and "result" not in line.lower():
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "result" in line.lower():
                    code += line
                    if int(d.findall(line)[0]) <= 90:
                        state = "Y"
                        explain += "패스워드 최대 사용기간이 90일 이하로 설정되어 있다.\n"
                    else:
                        state = "N"
                        explain += "패스워드 최대 사용기간이 90일 이하로 설정되어 있지 않다.\n"
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ###패스워드 최소 사용기간 설정
        elif "U-09" in num:
            n = 9
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None and "result" not in line.lower():
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "result" in line.lower():
                    code += line
                    if int(d.findall(line)[0]) >= 30:
                        state = "Y"
                        explain += "패스워드 최소 사용기간이 30일 이상으로 설정되어 있다.\n"
                    else:
                        state = "N"
                        explain += "패스워드 최소 사용기간이 짧게 설정되어 있다. (최소 30일)\n"
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ###불필요한 계정 제거
        elif "U-10" in num:
            n = 10
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "Shell" in line: 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                    
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if line.split(" ")[0] in ["adm", "lp", "sync", "shutdown", "halt", "news", "uucp", 
                                          "operator", "games", "gopher", "nfsnobody", "squid"]:
                    state = "C"
                    code += line
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "불필요하게 등록된 계정이 존재하지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "불필요하게 보이는 계정이 존재한다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "불필요한 계정이 존재한다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ### 관리자 그룹에 최소한의 계정 포함
        elif "U-11" in num:
            n = 11
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "group" in line.lower() and "members" in line.lower(): 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                    
                flag = 0
                l = line.lower().replace("root", "")
                for c in string.ascii_lowercase:
                    if (l.count(c) > 0):
                        flag = 1
                        break
                
                if flag == 0:
                    state = "Y"
                    explain += "관리자 그룹에 불필요한 계정이 등록되어 있지 않다."
                else:
                    state = "C"
                    explain += "관리자 그룹에 불필요해 보이는 계정이 등록되어 있다."
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "불필요한 계정이 존재한다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 계정이 존재하지 않는 GID 금지
        elif "U-12" in num:
            n = 12
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif ":" in line:
                    state = "C"
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "구성원이 없는 그룹이 존재하지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "구성원이 없는 그룹이 존재한다. 이 그룹이 사용되는 그룹인지 확인 필요\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "구성원이 없는 그룹이 존재한다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ###동일한 UID 점검
        elif "U-13" in num:
            n = 13
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            line = f.readline()
            code += "\n###" + line.split("[")[1][:-2] + "###\n"
            while True:
                line = f.readline()
                list_id = []
                
                if " END" in line: 
                    num = ""
                    break
                    
                if (":") in line:
                    if line.split(":")[1] in list_id:
                        state = "N"
                        code += line
                        break
                    else:
                        list_id.append(line.split(":")[1])
                    
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "중복된 UID가 존재하지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "확인 필요\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "중복된 UID가 존재한다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### U-14 사용자 shell 점검
        elif "U-14" in num:
            n = 14
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "Shell" in line: 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                    
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if line.split(" ")[0] in ["daemon", "bin", "sys", "adm", "listen",
                                          "nobody", "nobody4", "noaccess", "diag", "operator", "games", "gopher"]:
                    if "/sbin/nologin" not in line and "/bin/false" not in line:
                        state = "C"
                        code += line
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "로그인이 필요하지 않은 계정에 /bin/false(nologin) 쉘이 부여되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "로그인이 필요하지 않은 계정에 /bin/false(nologin) 쉘이 부여되어 있지 않다..\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "로그인이 필요하지 않은 계정에 /bin/false(nologin) 쉘이 부여되어 있지 않다..\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ### Session Timeout 설정
        elif "U-15" in num:
            n = 15
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            flag = 0
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "timeout" in line.lower():
                    code += line
                    if int(d.findall(line)[0]) <= 600:
                        explain += "Session Timeout이 600초 이하로 설정되어 있다.\n"
                    else:
                        state = "N"
                        explain += "Session Timeout이 600초 이하로 설정되어 있지 않다.\n"
                elif "export" in line.lower():
                    code += line
                    flag = 1
                elif "autologout" in line.lower():
                    code += line
                    if int(d.findall(line)[0]) <= 10:
                        explain += "자동 로그아웃 설정이 10분 이하로 설정되어 있다.\n"
                    else:
                        state = "N"
                        explain += "자동 로그아웃 설정이 10분 이하로 설정되어 있지 않다.\n"
            
            if state == "Y" and flag == 0:
                state = "N"
                explain += "export TMOUT 설정이 존재하지 않는다.\n"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### root 홈, 패스 디렉터리 권한 및 패스 설정
        elif "U-16" in num:
            n = 16
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    code += line
                    if line.count("./") > 0 or line.count("::/") > 0 or line.count(".:/") > 0:
                        state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "Path 환경변수에 '.'이 포함되지 않았다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Path 환경변수에 '.'이 포함되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)      
        
        ### 없네...;;
        #elif "U-17" in num:
        
        ### /etc/passwd 파일 소유자 및 권한 설정
        elif "U-18" in num:
            n = 18
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/etc/passwd" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/passwd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/passwd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/passwd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/passwd 파일의 소유자가 root가 아니다.\n"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/passwd 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /etc/shadow 파일 소유자 및 권한 설정
        elif "U-19" in num:
            n = 19
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/etc/shadow" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/shadow 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/shadow 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/shadow 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/shadow 파일의 소유자가 root가 아니다.\n"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/shadow 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /etc/hosts 파일 소유자 및 권한 설정
        elif "U-20" in num:
            n = 20
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/etc/hosts" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/hosts 파일의 소유자가 root가 아니다.\n"
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/hosts 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /etc/(x)inetd.conf 파일 소유자 및 권한 설정
        elif "U-21" in num:
            n = 21
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "xinetd" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/(x)inetd.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/(x)inetd.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/(x)inetd.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/(x)inetd.conf 파일의 소유자가 root가 아니다.\n"
                            
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/(x)inetd.conf 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /etc/syslog.conf 파일 소유자 및 권한 설정
        elif "U-22" in num:
            n = 22
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "syslog" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/syslog.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/syslog.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/syslog.conf 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] not in ["root", "bin", "sys"]:
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/syslog.conf 파일의 소유자가 root가 아니다.\n"
                            
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/syslog.conf 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /etc/services 파일 소유자 및 권한 설정
        elif "U-23" in num:
            n = 23
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "services" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/services 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/services 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/services 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] not in ["root", "bin", "sys"]:
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/services 파일의 소유자가 root가 아니다.\n"
                            
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "/etc/services 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### SUID, SGID, Sticky bit 설정파일 점검
        elif "U-24" in num:
            n = 24
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    if "s" in line.split(" ")[0].lower():
                        code += line
                        state = "N"
                    if "t" in line.split(" ")[0].lower():
                        code += line
                        state = "N"
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "주요 실행파일의 권한에 SUID와 SGID에 대한 설정이 부여되어 있지 않다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "주요 실행파일의 권한에 SUID와 SGID에 대한 설정이 부여되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 사용자, 시스템 시작파일 및 환경파일 소유자 및 권한 설정
        elif "U-25" in num:
            n = 25
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    code += line
                    
                    if "home" in line:
                        if line.split("/")[2] != line.split(" ")[2] and line.split(" ")[2] != "root":
                            state = "N"
                            
                            if "지정되지 않았다." not in explain:
                                explain += "환경변수 파일 소유자가 root 또는 해당 계정으로 지정되지 않았다.\n"
                    else:
                        if line.split(" ")[2] != "root":
                            state = "N"
                            
                            if "지정되지 않았다." not in explain:
                                explain += "환경변수 파일 소유자가 root로 지정되지 않았다.\n"
                    
                    if line[8] == 'w':
                        state = "N"
                        
                        if "쓰기 권한" not in explain:
                            explain += "외부 사용자에게 쓰기 권한이 부여되어 있다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9,color="0070C0")
                explain += "시작파일 및 환경파일의 소유자 / 권한이 올바르게 설정되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 없다.,..
        #elif "U-26" in num:
        
        ### /dev에 존재하지 않는 device 파일 점검
        elif "U-27" in num:
            n = 27
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    code += line
                    state = "N"
                    
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "존재하지 않은 device 파일이 없다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "존재하지 않은 device 파일이 제거되지 않았다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### /.rhosts, hosts.equiv 사용 금지
        elif "U-28" in num:
            n = 28
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "hosts.equiv" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split("/")[2] != line.split(" ")[2] and line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일의 소유자가 root 또는 파일 소유자가 아니다.\n"
                    
                    if "+" in line:
                        state = "N"
                        explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일 설정에 '+' 설정이 존재한다."
                                 
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size=9, color="0070C0")
                explain += "/etc/hosts.equiv 및 $HOME/.rhosts 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 접속 IP 및 포트 제한
        elif "U-29" in num:
            n = 29
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if "/etc/hosts.deny" in line and "-" in line: 
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                    
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if "all:all" in line:
                    code += line
                    if comment.match(line) != None:
                        explain += "포트 제한 설정이 주석처리 되어 있다."
                    else:
                        state = "Y"
                        explain += "포트 제한을 설정했다."
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                if explain == "":
                    explain += "포트 제한 설정이 없다."
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### hosts.lpd 파일 소유자 및 권한 설정
        elif "U-30" in num:
            n=30
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "hosts.lpd" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "hosts.lpd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "hosts.lpd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "hosts.lpd 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if line.split(" ")[2] == "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "hosts.lpd 파일의 소유자가 root가 아니다.\n"
                            
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size=9, color="0070C0")
                explain += "hosts.lpd 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### NIS 서비스 비활성화
        elif "U-31" in num:
            n=31
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "비실행중" in line:
                    state = "Y"
                    
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "불필요한 NIS 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "불필요한 NIS 서비스가 실행 중이다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### UMASK 설정 관리
        elif "U-32" in num:
            n=32
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif bool(re.search(r'\d', line)) == True:
                    code += line
                    if int(d.findall(line)[0]) > 22:
                        state = "N"
                        explain += "UMASK 값이 022 이하로 설정되어 있지 않다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "UMASK 값이 022 이하로 설정되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)         
        
        ### 홈디렉토리 소유자 및 권한 설정
        elif "U-33" in num:
            n=33
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line and "파일" not in line:
                    
                    if "home" in line:
                        if line.split("/")[2][:-1] != line.split(" ")[2]:
                            state = "N"
                            code += line
                            
                            if "아니다" not in explain:
                                explain += "홈 디렉토리 소유자가 해당 계정이 아니다.\n"
                    elif line.split(" ")[2] != "root":
                        code += line
                        state = "N"
                        
                        if "아니다" not in explain:
                                explain += "홈 디렉토리 소유자가 해당 계정이 아니다.\n"
                    
                    if line[8] == 'w':
                        state = "N"
                        code += line
                        
                        if "쓰기 권한" not in explain:
                            explain += "외부 사용자에게 쓰기 권한이 부여되어 있다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "홈 디렉토리의 소유자 / 권한이 올바르게 설정되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 홈디렉토리로 지정한 디렉토리의 존재 관리
        elif "U-34" in num:
            n = 34
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "|  " in line:
                    state = "N"
                    code += line
                elif "없음" in line:
                    state = "N"
                    code += line
                
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "홈 디렉토리가 존재하지 않는 계정이 발견되지 않았다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "홈 디렉토리가 존재하지 않는 계정이 발견되었다."
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 숨겨진 파일 및 디렉토리 검색 및 제거
        elif "U-35" in num:
            n = 35
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    if line.split("/")[-1][0] == ".":
                        state = "C"
                        code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "숨김 파일이 발견되지 않았다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "숨김 파일 확인, 삭제 여부 인터뷰 \n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "불필요한 숨김 파일을 발견하였다."
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Finger 서비스 비활성화
        elif "U-36" in num:
            n=36
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "/" in line:
                    if "/fingered fingerd" in line and comment.match(line) == None:
                        state = "C"
                        code += line
                        explain += "Finger 서비스 사용 중이다."
                elif "없" in line or "않" in line:
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "Finger 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Finger 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Anonymous FTP 비활성화
        elif "U-37" in num:
            n = 37
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    state = "Y"
                    code += line      
                    explain += "Anonymous FTP 미실행 중\n"
                elif "ftp:x" in line.lower():
                    code += line
                    if state == "Y":
                        explain += "FTP 계정 삭제 권고\n"
                    else:
                        explain += "FTP 계정을 삭제해야 한다.\n"
                elif "anonymous_enable" in line:
                    code += line
                    if line.split("=")[1] != "NO" or comment.match(line) != None:
                        if state == "N":
                            explain += "Anonymous FTP 접속 제한 설정이 필요하다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### r 계열 서비스 비활성화
        elif "U-38" in num:
            n = 38
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "없" in line or "않" in line:
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "r 계열 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080") 
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "r 계열 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### cron 파일 소유자 및 권한 설정
        elif "U-39" in num:
            n = 39
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "없" in line:
                    code += line
                    state = "NA"
                elif "cron" in line:
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "cron 관련 파일이 640 이상이다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "cron 관련 파일이 640 이상이다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "cron 관련 파일이 640 이상이다.\n"
                    
                    if line.split(" ")[2] != "root":
                        state = "N"
                        if "아니다" not in explain:
                            explain += "cron 관련 파일의 소유자가 root가 아니다.\n"
                                 
                elif "해당 파일이" in line:
                    code += line
            
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "cron 관련 파일이 적절한 권한을 가지고 있다.\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "cron 관련 파일이 존재하지 않습니다."
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Dos 공격에 취약한 서비스 비활성화
        elif "U-40" in num:
            n = 40
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "echo" in line or "discard" in line or "daytime" in line or "chargen" in line:
                    code += line
                    state = "N"
                elif "없" in line or "않" in line:
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "Dos 공격에 취약한 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080") 
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Dos 공격에 취약한 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### NFS 서비스 비활성화
        elif "U-41" in num:
            n = 41
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "root" in line:
                    code += line
                    state = "C"
                elif "없" in line or "않" in line:
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "NFS 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "NFS 서비스가 활성화 되어 있다. (불필요한 서비스인지 확인 필요)\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "NFS 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ###NFS 접근 통제
        elif "U-42" in num:
            n=42
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "share -F nfs -o" in line:
                    code += line
                    state = "N"
                elif "없" in line or "않" in line:
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "NFS 서비스 사용 안함\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")  
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "everyone 공유를 제한하지 않음\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
            
        ### automount 제거
        elif "U-43" in num:
            n = 43
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    state = "Y"
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "automountd 서비스가 비활성화\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "automountd 서비스가 활성화\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### RPC 서비스 구동 점검
        elif "U-44" in num:
            n=44
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "비 활성화" in line:
                    state = "Y"
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "RPC 서비스가 비활성화\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")  
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "RPC 서비스가 활성화\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### NIS, NIS+ 점검
        elif "U-45" in num:
            n=45
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "비실행" in line:
                    state = "Y"
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "NIS, NIS+ 서비스가 비활성화\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")  
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "NIS, NIS+ 서비스가 활성화\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### tftp, talk 서비스 비활성화
        elif "U-46" in num:
            n = 46
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "없" in line or "않" in line: 
                    state = "Y"
                    code += line
                        
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "tftp, talk, ntalk 서비스가 비활성화\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080") 
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "tftp, talk, ntalk 서비스가 활성화\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Sendmail 버전 점검
        elif "U-47" in num:
            n = 47
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line :
                    state = "NA"
                elif "." in line:
                    code += line
                    
                    if int(d.findall(line)[0]) < 8:
                        state = "N"
                    elif int(d.findall(line)[1]) < 13:
                        state = "N"
                    elif int(d.findall(line)[2]) < 8:
                        state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "최신 Sendmail 서비스 사용\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "Sendmail 서비스 사용하지 않음\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Sendmail 서비스 업데이트 필요\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 스팸 메일 릴레이 제한
        elif "U-48" in num:
            n = 48
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    code += line
                    state = "NA"
                elif "550 Relaying denied" in line:
                    code += line
                    if comment.match(line) == None:
                        state = "Y"
                    else:
                        explain += "sendmail 접근 제한 설정 주석 제거 필요\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "릴레이 제한 설정 존재\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "Sendmail 서비스 사용하지 않음"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                if explain == "" :
                    explain += "Sendmail 접근 제한 설정 필요\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### U-49 일반사용자의 Sendmail 실행 방지
        elif "U-49" in num:
            n = 49
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    code += line
                    state = "NA"
                elif "PrivacyOptions" in line:
                    code += line
                    if "restrictqrun" in line:
                        state = "Y"
                    else:
                        explain += "restrictqrun 옵션 추가해야 한다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "일반 사용자의 Sendmail 실행 방지 설정 존재\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "Sendmail 서비스 사용하지 않음"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                if explain == "" :
                    explain += "일반 사용자의 Sendmail 실행 방지 설정 필요\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### DNS 보안 버전 패치
        elif "U-50" in num:
            n=50
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    code += line
                    state = "NA"
                    explain += "DNS 사용하지 않음\n"
                elif "." in line:
                    code += line
                    
                    if int(d.findall(line)[0]) in [8, 9]:
                        state = "Y"
                    elif int(d.findall(line)[1]) in [3, 4]:
                        state = "Y"
                    elif int(d.findall(line)[2]) < 8:
                        state = "Y"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                if explain == "":
                    explain += "DNS 주기적으로 패치 적용\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "취약점이 존재하는 DNS 사용\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### DNS Zone Transfer 설정
        elif "U-51" in num:
            n=51
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    code += line
                    state = "NA"
                    explain += "DNS 사용하지 않음\n"
                elif "allow-transfer" in line and comment.match(line) == None:
                    code += line
                    state = "Y"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "Zone Transfer를 허가된 사용자에게만 허용\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Zone Transfer를 모든 사용자에게만 허용\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 디렉터리 리스팅 제거
        elif "U-52" in num:
            n = 52
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line:
                    code += line
                    state = "NA"
                    explain += "Apache 사용하지 않음\n"
                elif "options" in line.lower():
                    code += line
                    
                    if "indexes" in line.lower():
                        state = "N"
                        explain += "Option 지시자에 Indexes 옵션이 존재하기 때문에 취약\n"
                    else:
                        state = "Y"
                        explain += "Option 지시자에 Indexes 옵션이 존재하지 않기 때문에 양호\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 웹 프로세스 권한 제한
        elif "U-53" in num:
            n=53
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "root" in line.lower():
                    state = "N"
                    code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "Apache 데몬이 root 권한으로 구동되지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "Apache 데몬이 root 권한으로 구동된다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 상위 디렉터리 접근 금지
        elif "U-54" in num:
            n=54
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "allowoverride" in line.lower():
                    if "None" in line:
                        state = "N"
                        code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "AllowOverride 설정이 None으로 되어 있지 않기 때문에 양호.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "AllowOverride 설정이 None으로 되어 있기 때문에 취약\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 불필요한 파일 제거
        elif "U-55" in num:
            n=55
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "-" in line.lower():
                    state = "C"
                    code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "불필요한 파일이 존재하지 않는다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "불필요해 보이는 파일이 존재한다\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "불필요한 파일이 존재한다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 링크 사용금지
        elif "U-56" in num:
            n=56
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "options" in line.lower():
                    
                    if "FollowSymLinks" in line:
                        state = "N"
                        code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "심볼릭 링크 사용을 제한했다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "심볼릭 링크 사용을 제한하지 않았다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ###Apache 파일 업로드 및 다운로드 제한
        elif "U-57" in num:
            n=57
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "LimitRequestBody" in line:
                    code += line
                    
                    if d.findall(line)[0] > 5000000:
                        state = "N"
                        explain += "업로드 및 다운로드 파일이 5M를 넘었다.\n"
                elif "설정" in line:
                    code += line
                    state = "N"
                    explain += "파일 업로드 및 다운로드 제한 설정이 없다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "파일 업로드 및 다운로드를 제한했다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 웹 서비스 영역의 분리
        elif "U-58" in num:
            n=58
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "documentroot" in line.lower():
                    code += line
                    
                    l = l + "/" + line.split("/")[-1]
                    if "/usr/local/apache/htdocs" in line or "/usr/local/apache2/htdocs" in line or "/var/www/html" in line:
                        state = "N"
                        explain += "DocumentRoot가 별도의 디렉터리로 지정되지 않았다.\n"
                    elif "/usr/local/apache/htdocs" in l or "/usr/local/apache2/htdocs" in l or "/var/www/html" in l:
                        state = "N"
                        explain += "DocumentRoot가 별도의 디렉터리로 지정되지 않았다.\n"
                elif "httpd_root" in line.lower():
                    code += line                    
                    l = line.split(" ")[-1]

                elif "분리되어 있습니다." in line:
                    code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "DocumentRoot를 별도의 디렉토리로 지정했다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 시스템 보안 접근 설정
        elif "U-59" in num:
            n=59
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "telnet" in line.lower() and "참고" not in line:
                    code += line
                    state = "N"
                    
                    if "Telnet" not in explain:
                        explain += "Telnet 프로토콜이 사용 중이다."
                elif "ftp" in line.lower() and "참고" not in line:
                    code += line
                    state = "N"
                    
                    if "FTP" not in explain:
                        explain += "FTP 프로토콜이 사용 중이다."
                elif "sshd" in line:
                    code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "원격 접속 시 ssh 프로토콜을 사용한다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### FTP 서비스 구동 점검
        elif "U-60" in num:
            n=60
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                elif "ftp" in line.lower():
                    code += line
                    
                    if (comment.match(line) != None):
                        state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "FTP 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "FTP 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### FTP 계정 shell 제한
        elif "U-61" in num:
            n=61
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                    state = "NA"
                elif "ftp" in line.lower():
                    code += line
                    
                    if "/sbin/false" in line:
                        state = "Y"
                        explain += "ftp 계정에 /bin/false 쉘이 부여되어 있다.\n"
                    else:
                        state = "N"
                        explain += "ftp 계정에 /bin/false 쉘이 부여되지 않았다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "FTP 서비스가 비활성화 되어 있다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Ftpusers 파일 소유자 및 권한 설정
        elif "U-62" in num:
            n=62
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                    state = "NA"
                elif "ftpusers" in line.lower():
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "ftpusers 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "ftpusers 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "ftpusers 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if "root" not in line.split(" ")[2]:
                        state = "N"
                        if "아니다" not in explain:
                            explain += "ftpusers 파일의 소유자가 root가 아니다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "FTP 서비스가 비활성화 되어 있다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Ftpusers 파일 설정
        elif "U-63" in num:
            n=63
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                l = ""
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                    state = "NA"
                elif "root" in line.lower():
                    code += line
                    
                    if comment.match(line) != None:
                        state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "FTP 서비스가 비활성화 되어 있고, FTP 서비스가 활성화 시 root 계정 접속을 허용했다.\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "FTP 서비스가 비활성화 되어 있다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "FTP 서비사가 활성화되어 있고, root 계정 접속을 허용했다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### at 파일 소유자 및 권한 설정
        elif "U-64" in num:
            n=64
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            flag = 0
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "없" in line and flag != 0:
                    state = "NA"
                elif "at" in line.lower() and "파일" not in line:
                    flag = 1
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "at 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "at 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 0:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "at 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if "root" not in line.split(" ")[3]:
                        state = "N"
                        if "아니다" not in explain:
                            explain += "at 파일의 소유자가 root가 아니다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "at 파일의 소유자가 root이고, 적절한 권한을 가지고 있다.\n"
            elif state == "NA":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "at 접근제어 파일이 존재하지 않습니다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### SNMP 서비스 구동 점검
        elif "U-65" in num:
            n=65
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                elif "snmp" in line.lower():
                    code += line
                    state = "N"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "SNMP 서비스가 비활성화 되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "SNMP 서비스가 활성화 되어 있다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### SNMP 서비스 커뮤니티스트링의 복잡성 설정
        elif "U-66" in num:
            n=66
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                    state = "C"
                elif "com2sec" in line.lower():
                    code += line
                    
                    if "public" in line:
                        state = "N"
                        
                        if "public" not in line:
                            explain += "SNMP Community 이름이 public이다.\n"
                    elif "private" in line:
                        state = "N"
                        
                        if "private" not in line:
                            explain += "SNMP Community 이름이 private이다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "SNMP Community 이름이 디폴트 값이 아니다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "SNMP 서비스가 비활성화 되어 있다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ##로그온 시 경고 메시지 제공
        elif "U-67" in num:
            n=67
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            flag = 0
            while True:
                line = f.readline()
                if "2. SSH" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "banner" in line.lower():
                    code += line
                    flag= 1
            
            if flag == 0:
                state = "N"
                explain += "TCP에 로그온 메시지가 설정되어 있지 않다.\n"
            while True:
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "greetingmessage" in line.lower():
                    code += line
                    flag= 1
                line = f.readline()
            
            if flag == 0:
                state = "N"
                explain += "SMTP에 로그온 메시지가 설정되어 있지 않다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "로그온 메시지가 설정되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### NFS 설정파일 접근권한
        elif "U-68" in num:
            n=68
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "없" in line:
                    code += line
                    state = "C"
                elif "at" in line.lower() and "파일" not in line:
                    flag = 1
                    code += line
                    total = 0
                    
                    for i in range(1, 10):
                        if line[i] == 'r':
                            total += 4
                        elif line[i] == 'w':
                            total += 2
                        elif line[i] == 'x':
                            total += 1
                        
                        if i % 3 == 0 and i != 9:
                            total *= 10
                    
                    if int(total / 100 ) > 6:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "NFS 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if int(total / 10) % 10 > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "NFS 파일이 적절한 권한을 가지고 있지 않다.\n"
                    if (total % 10) > 4:
                        state = "N"
                        if "있지 않다" not in explain:
                            explain += "NFS 파일이 적절한 권한을 가지고 있지 않다.\n"
                    
                    if "root" not in line.split(" ")[3]:
                        state = "N"
                        if "아니다" not in explain:
                            explain += "NFS 파일의 소유자가 root가 아니다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "NFS 파일의 소유자가 root이고, 적절한 권한을 가지고 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "NFS 파일이 존재하지 않습니다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### expn, vrfy 명령어 제한
        elif "U-69" in num:
            n=69
            state = "Y"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "미실행" in line.lower():
                    code += line
                    state = "Y"
                    explain += "SMTP 서비스 미사용"
                elif "privacyoptions" in line.lower():
                    code += line
                    
                    if "noexpn" not in line.lower() or "novrfy" not in line.lower():
                        state = "N"
                    else: 
                        if "SMTP" not in explain:
                            explain += "SMTP 서비스에 noexpn, novrfy 옵션이 설정되어 있다.\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "N/A"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "SNMP 서비스가 비활성화 되어 있다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "SMTP 서비스에 noexpn, novrfy 옵션이 설정되어 있지 않다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### Apache 웹서비스 정보 숨김
        elif "U-70" in num:
            n=70
            state = "N"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif "servertokens" in line.lower():
                    code += line
                    if "prob" in line.lower():
                        state = "Y"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
                explain += "ServerTokens 지시자에 Prob 옵션이 설정되어 있다.\n"
            elif state == "C":
                s['E' + str(n + 6)] = "Error"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
                explain += "ServerTokens 지시자에 Prob 옵션이 설정되어 있지 않다.\n"
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        # 최신 보안패치 및 벤더 권고사항 적용
        elif "U-71" in num:
            n=71
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "최신 보안패치 적용에 대한 인터뷰가 필요하다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        ### 로그의 정기적 검토 및 보고
        elif "U-72" in num:
            n=72
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "로그의 정기적 검토 및 보고에 대한 인터뷰가 필요하다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
        
        elif "U-73" in num:
            n=73
            state = "C"
            explain="[현재 상태]\n"
            code = ""
            
            u75 = re.compile('\s+')
            while True:
                line = f.readline()
                if " END" in line: 
                    num = ""
                    break
                
                if subt.match(line) != None:
                    code += "\n###" + line.split("[")[1][:-2] + "###\n"
                elif comment.match(line) == None and u75.match(line) == None:
                    code += line
                        
            if state == "Y":
                s['E' + str(n + 6)] = "양호"
                s['E' + str(n + 6)].font = Font(size = 9, color="0070C0")
            elif state == "C":
                s['E' + str(n + 6)] = "인터뷰 필요"
                s['E' + str(n + 6)].font = Font(size = 9, color="808080")
                explain += "로그의 정책에 대한 인터뷰가 필요하다.\n"
            else:
                s['E' + str(n + 6)] = "취약"
                s['E' + str(n + 6)].font = Font(size = 9, color="FF0000")
            s['F' + str(n + 6)] = explain + code
            result.append(state)
    f.close()
    
    print("################ [" + server_name + "] 점검 종료 ################ ")

    for i in range(73):
        if (i + 1) < 10:
            s['B' + str(i + 7)] = "U-" + "0" + str(i + 1)
        else:
            s['B' + str(i + 7)] = "U-" + str(i + 1)
        
        s['C' + str(i + 7)] = checklist[i]
        s['D' + str(i + 7)] = important[i]
        
        for j in range(8):
            if j == 3:
                if important[i] == "상":
                    s.cell(row=i+7, column=j+1).font = Font(size = 9, color="FF0000")
                elif important[i] == "중":
                    s.cell(row=i+7, column=j+1).font = Font(size = 9, color="0070C0")
                else:
                    s.cell(row=i+7, column=j+1).font = Font(size = 9, color="000000")
            elif j != 4:
                s.cell(row = i + 7, column = j + 1).font = Font(size = 9)
                s.cell(row=i+7, column=j+1).alignment = Alignment(horizontal='center')
            
            if j in [0, 1, 3]:
                s.cell(row=i+7, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
            elif j != 4:
                s.cell(row=i+7, column=j+1).alignment = Alignment(horizontal='left', vertical='center')
            if (i + 7) >= 22 and (i + 7) <= 41:
                s.cell(row=i+7,column=j+1).fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')
            elif (i+7) == 77:
                s.cell(row=i+7,column=j+1).fill = PatternFill(start_color='DCE6F2', end_color='DCE6F2', fill_type='solid')
            
            s.cell(row=i+7, column=j+1).border = Border(left=Side(style='dotted'), 
                                                            top=Side(style='dotted'),
                                                            right=Side(style='dotted'), 
                                                            bottom=Side(style='dotted'))
        
    s.merge_cells("A7:A21")
    s['A7'] = "계정\n관리"
    s.merge_cells("A22:A41")
    s['A22'] = "접근\n제어"
    s.merge_cells("A42:A76")
    s['A42'] = "서비스\n관리"
    s['A77'] = "패치\n관리"
    s.merge_cells("A78:A79")
    s['A78'] = "로그\n관리"
    

wb.save("result/Linux 서버 취약점분석.xlsx")